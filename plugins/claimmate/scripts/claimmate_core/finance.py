from .base import *
from .documents import *
from .model import *
from .intake import *
from .requirements import *
from .policy import *

def find_review_document(state: dict[str, Any], query: str) -> tuple[str, dict[str, Any]]:
    normalized = normalize(query)
    matches = [
        (digest, record)
        for digest, record in state.get("documents", {}).items()
        if record.get("status") == "review"
        and (
            digest.startswith(query.lower())
            or normalized == normalize(str(record.get("original_name", "")))
            or normalized == normalize(str(record.get("current_path", "")))
        )
    ]
    if not matches:
        raise SystemExit(f"没有找到待核对文件：{query}")
    if len(matches) > 1:
        raise SystemExit(f"匹配到多个待核对文件，请使用 SHA-256 前缀指定：{query}")
    return matches[0]


@locked_command
def command_resolve(args: argparse.Namespace) -> None:
    """Apply a high-confidence document decision made by ClaimMate itself."""
    root = Path(args.folder).expanduser().resolve()
    registry, _ = load_project_registry(root)
    require_named_project(registry, args.project, "核对")
    state, config = load_workspace(root, project=args.project)
    prepare_workspace(root, state, config)
    digest, record = find_review_document(state, args.file)
    source = absolute(root, record["current_path"])
    if not source.exists():
        raise SystemExit(f"待核对文件不存在：{record['current_path']}")

    item = analyze(source, config, record.get("original_name"))
    item["sha256"] = digest
    item["added_at"] = record.get("added_at")
    item["from_observation"] = True
    if record.get("decision_source") == MODEL_DECISION_SOURCE:
        for key in (
            "role", "category", "merchant", "amount", "expense_label", "expense_type_key",
            "material_type", "applicable_requirement_ids", "assessed_condition_rule_ids",
            "date_tokens", "reference_tokens"
        ):
            if key in record:
                item[key] = record.get(key)
        item["role_confidence"] = record.get("role_confidence")
        item["category_confidence"] = record.get("category_confidence")
        item["decision_source"] = record.get("decision_source")
        item["model_provider"] = record.get("model_provider")
        item["model_confidence"] = record.get("model_confidence")
        item["model_expense_key"] = record.get("model_expense_key")
    item["inference_evidence"] = [
        *record.get("inference_evidence", []),
        "ClaimMate 文档复核确认",
    ]
    if args.role:
        item["role"] = args.role
        item["material_type"] = {
            "invoice": "发票",
            "payment": "付款记录",
            "supporting": "补充材料",
        }.get(args.role)
        item["role_confidence"] = "confirmed"
    if args.category:
        resolved_category = category_key(args.category, config)
        if not resolved_category:
            raise SystemExit(f"未知费用类型：{args.category}")
        item["category"] = resolved_category
        item["category_confidence"] = "confirmed"
    if args.expense_name:
        item["expense_label"] = safe_name(args.expense_name, "待识别费用")

    preferred_expense_id = args.expense_id
    if preferred_expense_id:
        expense = state.get("expenses", {}).get(preferred_expense_id)
        if not expense:
            raise SystemExit(f"费用编号不存在：{preferred_expense_id}")
        item["category"] = expense.get("category")
        item["expense_label"] = expense.get("label") or item.get("expense_label")
        item_amount = as_amount(item.get("amount"))
        expense_amount = as_amount(expense.get("amount"))
        if item_amount is not None and expense_amount is not None and item_amount != expense_amount:
            raise SystemExit("文件金额与指定费用金额冲突，ClaimMate 不会自动覆盖。")
    else:
        preferred_expense_id = preferred_model_expense(item, state, {})

    if item.get("role") not in {"invoice", "payment", "supporting"}:
        raise SystemExit("仍无法确定文件是发票、付款记录还是补充材料。")
    if item.get("category") not in config.get("categories", {}):
        if preferred_expense_id:
            item["category"] = state["expenses"][preferred_expense_id]["category"]
        else:
            raise SystemExit("仍无法确定费用类型。")
    if item["role"] == "supporting" and not preferred_expense_id:
        raise SystemExit("补充材料尚不能唯一对应到某笔费用。")

    before_state = copy.deepcopy(state)
    operations: list[dict[str, str]] = []
    organize_item(
        root,
        item,
        state,
        config,
        operations,
        False,
        preferred_expense_id=preferred_expense_id,
    )
    if state["documents"].get(digest, {}).get("status") != "organized":
        raise SystemExit("该文件仍无法安全归类。")
    state["status"] = "collecting"
    state.pop("ready_at", None)
    save_state(root, state)
    results = {"resolved": 1, "source": record.get("original_name")}
    record_transaction(root, before_state, state, operations, results)
    resolved = state["documents"][digest]
    print(
        f"已根据材料线索确认：{resolved['original_name']} → "
        f"{resolved['expense_id']} / {role_label(resolved['role'])}。"
    )


@locked_command
def command_ready(args: argparse.Namespace) -> None:
    root = Path(args.folder).expanduser().resolve()
    registry, config = load_project_registry(root)
    require_setup_ready(root, config)
    require_named_project(registry, args.project, "提交")
    state = select_project(registry, args.project, include_archived=False)
    prepare_workspace(root, state, config)
    if state.get("status") == "archived":
        raise SystemExit("该报销事项已经归档。")

    before_registry = copy.deepcopy(registry)
    process_routed_inputs(
        root,
        registry,
        config,
        False,
        before_registry,
        force_model_revisit=True,
    )
    summary = build_summary(state)
    review_documents = [
        item for item in state.get("documents", {}).values() if item.get("status") == "review"
    ]
    missing_expenses = [
        (expense_id, expense, expense_status(expense, state))
        for expense_id, expense in sorted(state.get("expenses", {}).items())
        if not expense_status(expense, state)["complete"]
    ]
    duplicates = state.get("duplicates", [])
    unassigned_candidates = unassigned_candidates_for_project(registry, state)
    requirement_catalog = load_requirement_catalog(root, persist_snapshot=True)
    requirements_review, finance_feedback_review = evaluate_effective_requirements(
        root, state, requirement_catalog
    )
    state["requirements_review"] = requirements_review
    requirements_gaps = requirements_review.get("gaps", [])
    workbook_requirement_gaps = [
        item for item in requirements_gaps
        if not str(item.get("source") or "").startswith("finance-feedback:")
    ]
    state["finance_feedback_review"] = finance_feedback_review
    finance_feedback_gaps = finance_feedback_review.get("gaps", [])
    recipient = claimant_name(config)
    ready = (
        bool(state.get("expenses"))
        and bool(recipient)
        and not review_documents
        and not missing_expenses
        and not duplicates
        and not unassigned_candidates
        and not requirements_gaps
    )
    state["readiness_checked_at"] = now_iso()
    if ready:
        state["status"] = "ready"
        state["ready_at"] = now_iso()
    else:
        state["status"] = "collecting"
        state.pop("ready_at", None)
    save_state(root, state)

    print(f"报销前核验：{state['case_name']}")
    requirements_by_expense = {
        str(item.get("expense_id")): item
        for item in requirements_review.get("expenses", [])
    }
    requirements_passed = sum(
        1 for item in requirements_by_expense.values() if item.get("passed")
    )
    print(
        f"已识别 {summary['expenses']} 笔费用，"
        f"其中 {requirements_passed} 笔满足工作簿中的报销要求。"
    )
    if state.get("expenses"):
        print("\n费用汇总：")
        for expense_id, expense in sorted(state["expenses"].items()):
            details = expense_status(expense, state)
            category = expense.get("label") or config["categories"].get(
                expense.get("category"), {}
            ).get("label", expense.get("category"))
            requirement_status = requirements_by_expense.get(expense_id, {})
            status = expense_state_text(details)
            if details.get("complete"):
                status = "满足报销要求" if requirement_status.get("passed") else "待补材料"
            print(
                f"- {expense_id} | {category} | {expense.get('merchant')} | "
                f"{amount_text(expense.get('amount'))} | {status}"
            )
    if finance_feedback_review.get("entries"):
        print("\n财务反馈审查依据：")
        for line in finance_feedback_review_lines(finance_feedback_review):
            print(line)
    if ready:
        generated = export_claim(root, state, config)
        save_state(root, state)
        print("\n核验通过：基础材料和已记录的财务反馈要求均已满足，已生成报销明细表，可以交给财务。")
        print(f"默认收款人：{recipient}")
        print("已生成：")
        for path in generated:
            print(f"- {relative(path, root)}")
        return

    if not recipient:
        print("\n缺少使用者姓名：请先设置，交付明细表需要默认收款人。")
    if not state.get("expenses"):
        print("\n缺失：尚未识别到可报销费用。")
    if review_documents:
        print("\n仍不确定的文件：")
        for item in review_documents:
            print(f"- {item.get('original_name')}：{item.get('review_reason', '线索不足')}")
    if missing_expenses:
        print("\n仍需处理的材料或金额：")
        for expense_id, expense, details in missing_expenses:
            category = expense.get("label") or config["categories"].get(
                expense.get("category"), {}
            ).get("label", expense.get("category"))
            print(
                f"- {expense_id} | {category} | {expense.get('merchant')} | "
                f"{amount_text(expense.get('amount'))}：{expense_state_text(details)}"
            )
    if duplicates:
        print("\n需要核对的重复文件：")
        for item in duplicates:
            print(f"- {item.get('original_name')}")
    if unassigned_candidates:
        print("\n大模型关联到本项目、但仍未确定归属的文件：")
        for item in unassigned_candidates:
            evidence = "、".join(item.get("inference_evidence", [])) or "大模型线索不足"
            print(f"- {item.get('original_name')}：{evidence}")
    if finance_feedback_gaps:
        print("\n依据历史财务反馈仍需补充：")
        for item in finance_feedback_gaps:
            print(
                f"- {item.get('expense_id')} | {item.get('merchant')} | "
                f"依据 {item.get('feedback_id')} 缺少{'、'.join(item.get('missing_evidence', []))}"
            )
    if workbook_requirement_gaps:
        print("\n依据报销要求仍需补充：")
        for item in workbook_requirement_gaps:
            print(f"- {item.get('expense_id')}：{'、'.join(item.get('missing', []))}")
    print("\n以上是当前仅剩的确认项；补充材料后，再说“报销要交给财务了”即可重新核验。")


def excel_column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def write_detail_xlsx_fallback(
    path: Path,
    state: dict[str, Any],
    headers: list[str],
    rows: list[list[Any]],
    default_recipient: str,
) -> None:
    """Write the delivery workbook with only the Python standard library."""
    header_row = 4
    data_start = header_row + 1
    data_end = header_row + len(rows)
    total_row = data_end + 1
    last_column = excel_column_name(len(headers))
    totals = [
        sum((as_amount(row[index]) or Decimal("0")) for row in rows)
        for index in (4, 5, 6)
    ]

    def inline_cell(reference: str, value: Any, style: int = 0) -> str:
        rendered = escape(str(value))
        style_attribute = f' s="{style}"' if style else ""
        return (
            f'<c r="{reference}"{style_attribute} t="inlineStr">'
            f'<is><t xml:space="preserve">{rendered}</t></is></c>'
        )

    def numeric_cell(reference: str, value: Decimal, style: int) -> str:
        return f'<c r="{reference}" s="{style}"><v>{format(value, "f")}</v></c>'

    sheet_rows = [
        f'<row r="1" ht="30" customHeight="1">'
        + inline_cell("A1", f"{state.get('case_name') or '报销项目'} 报销明细表", 1)
        + "</row>",
        '<row r="2">'
        + inline_cell("A2", "默认收款人", 2)
        + inline_cell("B2", default_recipient)
        + inline_cell("D2", "生成时间", 2)
        + inline_cell("E2", now_iso())
        + "</row>",
        '<row r="3"></row>',
        f'<row r="{header_row}">'
        + "".join(
            inline_cell(f"{excel_column_name(column)}{header_row}", value, 3)
            for column, value in enumerate(headers, start=1)
        )
        + "</row>",
    ]
    for row_index, row in enumerate(rows, start=data_start):
        alternate = (row_index - data_start) % 2 == 1
        cells: list[str] = []
        for column, value in enumerate(row, start=1):
            reference = f"{excel_column_name(column)}{row_index}"
            if column in {5, 6, 7}:
                amount = as_amount(value)
                if amount is not None:
                    cells.append(numeric_cell(reference, amount, 6 if alternate else 4))
                continue
            if column in {8, 9, 10}:
                cells.append(
                    numeric_cell(reference, Decimal(str(value or 0)), 12 if alternate else 11)
                )
                continue
            style = 10 if alternate and column == 12 else 9 if column == 12 else 5 if alternate else 11
            cells.append(inline_cell(reference, value if value is not None else "", style))
        sheet_rows.append(f'<row r="{row_index}">' + "".join(cells) + "</row>")
    total_cells: list[str] = []
    for column, total in zip(("E", "F", "G"), totals):
        formula = f"SUM({column}{data_start}:{column}{data_end})" if rows else "0"
        total_cells.append(
            f'<c r="{column}{total_row}" s="8"><f>{formula}</f><v>{format(total, "f")}</v></c>'
            if rows
            else numeric_cell(f"{column}{total_row}", Decimal("0"), 8)
        )
    sheet_rows.append(
        f'<row r="{total_row}">'
        + inline_cell(f"D{total_row}", "合计", 7)
        + "".join(total_cells)
        + "</row>"
    )

    widths = [18, 18, 18, 28, 18, 18, 18, 14, 18, 18, 24, 72]
    columns_xml = "".join(
        f'<col min="{index}" max="{index}" width="{width}" customWidth="1"/>'
        for index, width in enumerate(widths, start=1)
    )
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="A1:{last_column}{total_row}"/>'
        '<sheetViews><sheetView showGridLines="0" workbookViewId="0">'
        f'<pane ySplit="4" topLeftCell="A{data_start}" activePane="bottomLeft" state="frozen"/>'
        f'<selection pane="bottomLeft" activeCell="A{data_start}" sqref="A{data_start}"/>'
        '</sheetView></sheetViews><sheetFormatPr defaultRowHeight="15"/>'
        f'<cols>{columns_xml}</cols><sheetData>{"".join(sheet_rows)}</sheetData>'
        f'<mergeCells count="2"><mergeCell ref="A1:{last_column}1"/><mergeCell ref="E2:{last_column}2"/></mergeCells>'
        f'<autoFilter ref="A{header_row}:{last_column}{max(data_end, header_row)}"/>'
        '<pageMargins left="0.25" right="0.25" top="0.5" bottom="0.5" header="0.2" footer="0.2"/>'
        '<pageSetup orientation="landscape" fitToWidth="1" fitToHeight="0"/>'
        '</worksheet>'
    )
    styles_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <numFmts count="1"><numFmt numFmtId="164" formatCode="¥#,##0.00"/></numFmts>
  <fonts count="4">
    <font><sz val="11"/><name val="Calibri"/></font>
    <font><b/><color rgb="FF0F5B62"/><sz val="16"/><name val="Calibri"/></font>
    <font><b/><color rgb="FF0F5B62"/><sz val="11"/><name val="Calibri"/></font>
    <font><b/><color rgb="FFFFFFFF"/><sz val="11"/><name val="Calibri"/></font>
  </fonts>
  <fills count="5">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF0F5B62"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF147D86"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFEEF7F7"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="2">
    <border><left/><right/><top/><bottom/><diagonal/></border>
    <border><left/><right/><top/><bottom style="thin"><color rgb="FFD7E2E4"/></bottom><diagonal/></border>
  </borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="13">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>
    <xf numFmtId="0" fontId="2" fillId="0" borderId="0" xfId="0" applyFont="1"/>
    <xf numFmtId="0" fontId="3" fillId="3" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>
    <xf numFmtId="164" fontId="0" fillId="0" borderId="1" xfId="0" applyNumberFormat="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="0" fillId="4" borderId="1" xfId="0" applyFill="1" applyBorder="1"/>
    <xf numFmtId="164" fontId="0" fillId="4" borderId="1" xfId="0" applyNumberFormat="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="2" fillId="0" borderId="0" xfId="0" applyFont="1"/>
    <xf numFmtId="164" fontId="2" fillId="0" borderId="0" xfId="0" applyNumberFormat="1" applyFont="1"/>
    <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="0" fillId="4" borderId="1" xfId="0" applyFill="1" applyBorder="1" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"/>
    <xf numFmtId="0" fontId="0" fillId="4" borderId="1" xfId="0" applyFill="1" applyBorder="1"/>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>'''
    created = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>'''
    root_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>'''
    workbook_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <bookViews><workbookView xWindow="0" yWindow="0" windowWidth="24000" windowHeight="12000"/></bookViews>
  <sheets><sheet name="报销明细" sheetId="1" r:id="rId1"/></sheets>
  <calcPr calcId="191029" fullCalcOnLoad="1" forceFullCalc="1"/>
</workbook>'''
    workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>'''
    app_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>ClaimMate</Application><AppVersion>1.0</AppVersion>
</Properties>'''
    core_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:creator>ClaimMate</dc:creator><dc:title>{escape(str(state.get("case_name") or "报销项目"))} 报销明细表</dc:title>
  <dcterms:created xsi:type="dcterms:W3CDTF">{created}</dcterms:created><dcterms:modified xsi:type="dcterms:W3CDTF">{created}</dcterms:modified>
</cp:coreProperties>'''
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as workbook:
        workbook.writestr("[Content_Types].xml", content_types)
        workbook.writestr("_rels/.rels", root_rels)
        workbook.writestr("docProps/app.xml", app_xml)
        workbook.writestr("docProps/core.xml", core_xml)
        workbook.writestr("xl/workbook.xml", workbook_xml)
        workbook.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        workbook.writestr("xl/styles.xml", styles_xml)
        workbook.writestr("xl/worksheets/sheet1.xml", sheet_xml)


def export_claim(root: Path, state: dict[str, Any], config: dict[str, Any]) -> list[Path]:
    default_recipient = claimant_name(config)
    if not default_recipient:
        raise SystemExit("缺少使用者姓名，无法生成交付明细表。请先设置使用者姓名。")
    output = active_case_path(root, state)
    output.mkdir(parents=True, exist_ok=True)
    requirement_catalog = load_requirement_catalog(root, persist_snapshot=True)
    requirements_review, finance_feedback_review = evaluate_effective_requirements(
        root, state, requirement_catalog
    )
    state["requirements_review"] = requirements_review
    requirement_gaps_by_expense = {
        str(item.get("expense_id")): [str(value) for value in item.get("missing", [])]
        for item in requirements_review.get("gaps", [])
    }
    state["finance_feedback_review"] = finance_feedback_review
    headers = [
        "费用编号", "费用类型", "收款人", "商户/事项", "金额（元）",
        "付款合计（元）", "付款差额（元）", "发票数", "付款记录数",
        "补充材料数", "状态", "材料文件",
    ]
    rows: list[list[Any]] = []
    missing_lines = ["# 缺失材料清单", ""]
    for expense_id, expense in sorted(state.get("expenses", {}).items()):
        details = expense_status(expense, state)
        category = expense.get("label") or config["categories"].get(
            expense.get("category"), {}
        ).get("label", expense.get("category"))
        files = [
            state["documents"][digest]["current_path"]
            for digest in expense.get("documents", []) if digest in state["documents"]
        ]
        missing = list(details["missing"])
        missing.extend(requirement_gaps_by_expense.get(expense_id, []))
        missing = list(dict.fromkeys(missing))
        status = expense_state_text({"complete": not missing, "missing": missing})
        recipient = str(expense.get("recipient") or default_recipient).strip()
        reconciliation = details.get("amount_reconciliation", {})
        rows.append([
            expense_id, category, recipient, expense.get("merchant"), expense.get("amount") or "",
            reconciliation.get("payment_total") or "",
            reconciliation.get("difference") if reconciliation.get("difference") is not None else "",
            details["invoice_count"], details["payment_count"], details["supporting_count"],
            status, "；".join(files),
        ])
        if missing:
            missing_lines.append(
                f"- {expense_id} {expense.get('merchant')}："
                f"{expense_state_text({'complete': False, 'missing': missing})}"
            )
    review_documents = [
        item for item in state.get("documents", {}).values() if item.get("status") == "review"
    ]
    if review_documents:
        missing_lines.extend(["", "## 待确认文件"])
        missing_lines.extend(
            f"- {item.get('current_path')}：{item.get('review_reason', '待人工确认')}"
            for item in review_documents
        )
    if state.get("duplicates"):
        missing_lines.extend(["", "## 重复文件"])
        missing_lines.extend(f"- {item.get('current_path')}" for item in state["duplicates"])
    if len(missing_lines) == 2:
        missing_lines.append("没有缺失或待确认材料。")

    csv_path = output / "报销明细表.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)
    generated = [csv_path]
    xlsx_path = output / "报销明细表.xlsx"
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "报销明细"
        last_column = get_column_letter(len(headers))
        sheet.merge_cells(f"A1:{last_column}1")
        sheet["A1"] = f"{state.get('case_name') or '报销项目'} 报销明细表"
        sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="0F5B62")
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 30
        sheet["A2"] = "默认收款人"
        sheet["B2"] = default_recipient
        sheet["D2"] = "生成时间"
        sheet["E2"] = now_iso()
        sheet["A2"].font = Font(bold=True, color="0F5B62")
        sheet["D2"].font = Font(bold=True, color="0F5B62")
        header_row = 4
        for column, value in enumerate(headers, start=1):
            sheet.cell(row=header_row, column=column, value=value)
        for row_index, row in enumerate(rows, start=header_row + 1):
            for column, value in enumerate(row, start=1):
                if column in {5, 6, 7}:
                    amount = as_amount(value)
                    value = float(amount) if amount is not None else None
                sheet.cell(row=row_index, column=column, value=value)
        thin = Side(style="thin", color="D7E2E4")
        header_fill = PatternFill("solid", fgColor="147D86")
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        data_start = header_row + 1
        data_end = header_row + len(rows)
        for row_index in range(data_start, data_end + 1):
            if (row_index - data_start) % 2:
                for cell in sheet[row_index]:
                    cell.fill = PatternFill("solid", fgColor="EEF7F7")
            for cell in sheet[row_index]:
                cell.alignment = Alignment(vertical="top", wrap_text=cell.column == 12)
                cell.border = Border(bottom=thin)
            for column in (5, 6, 7):
                sheet.cell(row=row_index, column=column).number_format = '¥#,##0.00'
        total_row = data_end + 1
        sheet.cell(row=total_row, column=4, value="合计")
        sheet.cell(row=total_row, column=4).font = Font(bold=True, color="0F5B62")
        if rows:
            sheet.cell(row=total_row, column=5, value=f"=SUM(E{data_start}:E{data_end})")
            sheet.cell(row=total_row, column=6, value=f"=SUM(F{data_start}:F{data_end})")
            sheet.cell(row=total_row, column=7, value=f"=SUM(G{data_start}:G{data_end})")
        else:
            sheet.cell(row=total_row, column=5, value=0)
            sheet.cell(row=total_row, column=6, value=0)
            sheet.cell(row=total_row, column=7, value=0)
        for column in (5, 6, 7):
            sheet.cell(row=total_row, column=column).font = Font(bold=True, color="0F5B62")
            sheet.cell(row=total_row, column=column).number_format = '¥#,##0.00'
        sheet.freeze_panes = f"A{data_start}"
        widths = [14, 14, 14, 24, 14, 14, 14, 10, 13, 13, 22, 66]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.sheet_view.showGridLines = False
        sheet.auto_filter.ref = f"A{header_row}:{last_column}{max(data_end, header_row)}"
        workbook.save(xlsx_path)
    except ImportError:
        write_detail_xlsx_fallback(xlsx_path, state, headers, rows, default_recipient)
    generated.append(xlsx_path)
    missing_path = output / "缺失材料清单.md"
    missing_path.write_text("\n".join(missing_lines) + "\n", encoding="utf-8")
    generated.append(missing_path)
    if finance_feedback_review.get("entries"):
        feedback_lines = [
            "# 财务反馈审查依据",
            "",
            f"核验时间：{finance_feedback_review.get('checked_at')}",
            f"报销项目：{state.get('case_name')}",
            "",
        ]
        feedback_lines.extend(finance_feedback_review_lines(finance_feedback_review))
        if finance_feedback_review.get("gaps"):
            feedback_lines.extend(["", "## 待补充证据"])
            feedback_lines.extend(
                f"- {item.get('expense_id')} {item.get('merchant')}：依据 "
                f"{item.get('feedback_id')} 缺少{'、'.join(item.get('missing_evidence', []))}"
                for item in finance_feedback_review["gaps"]
            )
        else:
            feedback_lines.extend(["", "没有因已记录财务反馈产生的待补充证据。"])
        feedback_path = output / "财务反馈审查依据.md"
        feedback_path.write_text("\n".join(feedback_lines) + "\n", encoding="utf-8")
        generated.append(feedback_path)
    report_path = output / "处理报告.json"
    write_json(report_path, {
        "generated_at": now_iso(),
        "claimant": {"name": default_recipient},
        "summary": build_summary(state),
        "expenses": state.get("expenses", {}),
        "review_documents": review_documents,
        "duplicates": state.get("duplicates", []),
        "requirements_review": requirements_review,
        "finance_feedback_review": finance_feedback_review,
    })
    generated.append(report_path)
    return generated


@locked_command
def command_export(args: argparse.Namespace) -> None:
    root = Path(args.folder).expanduser().resolve()
    registry, config = load_project_registry(root)
    require_setup_ready(root, config)
    require_named_project(registry, args.project, "导出")
    state, config = load_workspace(root, project=args.project)
    prepare_workspace(root, state, config)
    generated = export_claim(root, state, config)
    save_state(root, state)
    print("已生成：")
    for path in generated:
        print(f"- {relative(path, root)}")


@locked_command
def command_undo(args: argparse.Namespace) -> None:
    root = Path(args.folder).expanduser().resolve()
    registry, config = load_project_registry(root)
    if registry.get("projects"):
        active_id = registry.get("active_project_id")
        state = registry["projects"].get(active_id) or next(iter(registry["projects"].values()))
        prepare_workspace(root, state, config)
    paths = metadata_paths(root)
    transactions = sorted(paths["transactions"].glob("TX-*.json"), reverse=True)
    transaction_path = next((path for path in transactions if not read_json(path).get("undone")), None)
    if transaction_path is None:
        raise SystemExit("没有可撤销的整理操作。")
    transaction = read_json(transaction_path)
    restored = 0
    for operation in reversed(transaction.get("operations", [])):
        source = absolute(root, operation["from"])
        destination = absolute(root, operation["to"])
        if not destination.exists():
            continue
        source = unique_destination(source)
        source.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(destination), str(source))
        restored += 1
    before_state = transaction.get("before_state")
    if before_state is None:
        paths["state"].unlink(missing_ok=True)
        paths["projects"].unlink(missing_ok=True)
    elif "projects" in before_state:
        save_project_registry(root, before_state)
    else:
        save_state(root, before_state)
    transaction["undone"] = True
    transaction["undone_at"] = now_iso()
    write_json(transaction_path, transaction)
    print(f"已撤销 {transaction['id']}，恢复 {restored} 个文件。")


@locked_command
def command_archive(args: argparse.Namespace) -> None:
    root = Path(args.folder).expanduser().resolve()
    registry, config = load_project_registry(root)
    require_setup_ready(root, config)
    require_named_project(registry, args.project, "归档")
    state, config = load_workspace(root, project=args.project)
    prepare_workspace(root, state, config)
    if state.get("status") == "archived":
        raise SystemExit("该报销事项已经归档。")
    summary = build_summary(state)
    requirement_catalog = load_requirement_catalog(root, persist_snapshot=True)
    requirements_review, finance_feedback_review = evaluate_effective_requirements(
        root, state, requirement_catalog
    )
    state["requirements_review"] = requirements_review
    state["finance_feedback_review"] = finance_feedback_review
    unresolved = (
        summary.get("amount_conflicts", 0) +
        summary["review_items"] + summary["duplicates"] +
        len(requirements_review.get("gaps", []))
    )
    if unresolved and not args.force:
        raise SystemExit(
            "尚有未解决材料，已停止归档。请先补齐或确认；只有明确接受这些问题时才使用 --force。"
        )
    export_claim(root, state, config)
    source_case = active_case_path(root, state)
    finished_case = root / FINISHED / state["case_name"]
    merge_or_rename_directory(source_case, finished_case)
    stage_mapping = {
        f"{ACTIVE}/{state['case_name']}": f"{FINISHED}/{state['case_name']}"
    }
    migrate_state_paths(state, stage_mapping, state["case_name"])
    rewrite_transactions(root, stage_mapping, state["case_name"])
    archive_path = finished_case / f"报销归档_{timestamp()}.zip"
    state["status"] = "archived"
    state["archived_at"] = now_iso()
    state["archive_path"] = relative(archive_path, root)
    save_state(root, state)
    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in finished_case.rglob("*"):
            if not path.is_file() or path == archive_path:
                continue
            if path.suffix.lower() == ".zip":
                continue
            archive.write(path, path.relative_to(finished_case).as_posix())
        archive.writestr(
            "manifest/project-state.json",
            json.dumps(state, ensure_ascii=False, indent=2) + "\n",
        )
        feedback_review = state.get("finance_feedback_review", {})
        requirements_review = state.get("requirements_review", {})
        archive.writestr(
            "manifest/requirements-review.json",
            json.dumps(requirements_review, ensure_ascii=False, indent=2) + "\n",
        )
        requirements_workbook = requirements_workbook_path(root)
        if requirements_workbook.is_file():
            archive.write(requirements_workbook, "manifest/报销要求.xlsx")
        if feedback_review.get("entries"):
            archive.writestr(
                "manifest/finance-feedback-review.json",
                json.dumps(feedback_review, ensure_ascii=False, indent=2) + "\n",
            )
            for entry in feedback_review["entries"]:
                source_file = entry.get("source_file")
                if not source_file:
                    continue
                source_path = absolute(root, source_file)
                if source_path.exists() and source_path.is_file():
                    archive.write(
                        source_path,
                        f"manifest/finance-feedback-sources/{source_path.name}",
                    )
        archive.write(metadata_paths(root)["config"], "manifest/config.json")
    print(f"归档完成：{relative(archive_path, root)}")
    if unresolved:
        print(f"注意：归档包含 {unresolved} 项未解决问题。")
